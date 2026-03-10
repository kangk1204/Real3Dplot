from __future__ import annotations

import argparse
import math
import random
import webbrowser
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence

import polars as pl
import xlrd
from openpyxl import load_workbook

from .html_builder import PLOTLY_CDN_URL, render_dashboard_html

NULL_MARKERS = ["", "NULL", "null", "NaN", "nan", "N/A", "n/a"]


@dataclass(slots=True)
class PreparedFrame:
    frame: pl.DataFrame
    total_rows: int
    sampled_rows: int
    sampling_note: str | None
    source_name: str
    sheet_name: str | None = None


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="3dplot-dashboard",
        description="Generate a lightweight interactive 3D plot HTML dashboard from Excel, CSV, or TSV input.",
    )
    parser.add_argument("input", type=Path, help="Path to the source file (.xlsx, .xls, .csv, .tsv).")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Path to the output HTML file. Defaults to <input-stem>.dashboard.html",
    )
    parser.add_argument("--sheet", help="Excel sheet name or zero-based sheet index.")
    parser.add_argument("--x", help="Initial x-axis column.")
    parser.add_argument("--y", help="Initial y-axis column.")
    parser.add_argument("--z", help="Initial z-axis column.")
    parser.add_argument("--color", help="Initial color mapping column.")
    parser.add_argument("--size", help="Initial size mapping column.")
    parser.add_argument("--label", help="Initial label/search column.")
    parser.add_argument("--delimiter", help="Explicit delimiter for text files.")
    parser.add_argument(
        "--max-points",
        type=int,
        default=100_000,
        help="Maximum number of rows embedded into the HTML dashboard. Default: 100000",
    )
    parser.add_argument("--seed", type=int, default=42, help="Sampling seed. Default: 42")
    parser.add_argument("--title", help="Dashboard title. Defaults to the input file stem.")
    parser.add_argument(
        "--plotly-url",
        default=PLOTLY_CDN_URL,
        help="Plotly.js source URL. Defaults to the Plotly CDN bundle.",
    )
    parser.add_argument(
        "--open",
        action="store_true",
        help="Open the generated HTML file in the system browser after creation.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    input_path = args.input.expanduser().resolve()
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    if args.max_points <= 0:
        raise SystemExit("--max-points must be greater than 0.")

    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else input_path.with_suffix("").with_name(f"{input_path.stem}.dashboard.html")
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    prepared = load_input_frame(
        input_path=input_path,
        sheet=args.sheet,
        delimiter=args.delimiter,
        max_points=args.max_points,
        seed=args.seed,
    )
    payload = build_dashboard_payload(prepared, args)
    title = args.title or input_path.stem.replace("_", " ").strip() or "3D Plot Dashboard"
    html = render_dashboard_html(payload=payload, title=title, plotly_url=args.plotly_url)
    output_path.write_text(html, encoding="utf-8")

    if args.open:
        webbrowser.open(output_path.as_uri())

    print(f"HTML dashboard created: {output_path}")
    print(f"Rows embedded: {prepared.sampled_rows:,} / {prepared.total_rows:,}")
    if prepared.sheet_name:
        print(f"Sheet: {prepared.sheet_name}")
    if prepared.sampling_note:
        print(f"Sampling: {prepared.sampling_note}")
    return 0


def load_input_frame(
    input_path: Path,
    sheet: str | None,
    delimiter: str | None,
    max_points: int,
    seed: int,
) -> PreparedFrame:
    suffix = input_path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        return _load_delimited_frame(input_path, delimiter=delimiter, max_points=max_points, seed=seed)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _load_openpyxl_frame(input_path, sheet=sheet, max_points=max_points, seed=seed)
    if suffix == ".xls":
        return _load_xls_frame(input_path, sheet=sheet, max_points=max_points, seed=seed)
    raise SystemExit(f"Unsupported input type: {input_path.suffix}")


def _load_delimited_frame(input_path: Path, delimiter: str | None, max_points: int, seed: int) -> PreparedFrame:
    separator = delimiter or _detect_separator(input_path)
    lazy_frame = pl.scan_csv(
        str(input_path),
        separator=separator,
        try_parse_dates=True,
        infer_schema_length=5_000,
        null_values=NULL_MARKERS,
        low_memory=True,
    )
    total_rows = int(lazy_frame.select(pl.len().alias("row_count")).collect().item())
    sampled = lazy_frame
    sampling_note = None

    if total_rows > max_points:
        step = math.ceil(total_rows / max_points)
        offset = seed % step
        sampled = (
            lazy_frame.with_row_index("__row_index")
            .filter((pl.col("__row_index") % step) == offset)
            .drop("__row_index")
        )
        sampling_note = f"Systematic row sampling every {step} rows with offset {offset}"

    frame = sampled.collect()
    if frame.height > max_points:
        frame = frame.sample(n=max_points, seed=seed, shuffle=True)
        sampling_note = f"{sampling_note or 'Systematic sampling'} + final cap to {max_points:,} rows"

    return PreparedFrame(
        frame=frame,
        total_rows=total_rows,
        sampled_rows=frame.height,
        sampling_note=sampling_note,
        source_name=input_path.name,
    )


def _load_openpyxl_frame(input_path: Path, sheet: str | None, max_points: int, seed: int) -> PreparedFrame:
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    sheet_name, worksheet = _select_openpyxl_sheet(workbook, sheet)
    randomizer = random.Random(seed)
    header: list[str] | None = None
    reservoir: list[list[Any]] = []
    total_rows = 0

    for raw_row in worksheet.iter_rows(values_only=True):
        values = [_normalize_python_value(cell) for cell in raw_row]
        if header is None:
            if _is_empty_row(values):
                continue
            header = _normalize_headers(values)
            continue
        normalized = _normalize_row_length(values, len(header))
        if _is_empty_row(normalized):
            continue
        total_rows += 1
        if len(reservoir) < max_points:
            reservoir.append(normalized)
        else:
            index = randomizer.randint(0, total_rows - 1)
            if index < max_points:
                reservoir[index] = normalized

    workbook.close()
    if header is None:
        raise SystemExit(f"No header row found in {input_path}")

    frame = pl.DataFrame(reservoir, schema=header, orient="row", strict=False)
    sampling_note = None
    if total_rows > max_points:
        sampling_note = f"Reservoir sampling capped at {max_points:,} rows with seed {seed}"

    return PreparedFrame(
        frame=frame,
        total_rows=total_rows,
        sampled_rows=frame.height,
        sampling_note=sampling_note,
        source_name=input_path.name,
        sheet_name=sheet_name,
    )


def _load_xls_frame(input_path: Path, sheet: str | None, max_points: int, seed: int) -> PreparedFrame:
    workbook = xlrd.open_workbook(str(input_path), on_demand=True)
    sheet_name, worksheet = _select_xlrd_sheet(workbook, sheet)
    randomizer = random.Random(seed)
    header: list[str] | None = None
    reservoir: list[list[Any]] = []
    total_rows = 0

    for row_idx in range(worksheet.nrows):
        values = [_normalize_xls_cell(worksheet.cell(row_idx, col_idx), workbook.datemode) for col_idx in range(worksheet.ncols)]
        if header is None:
            if _is_empty_row(values):
                continue
            header = _normalize_headers(values)
            continue
        normalized = _normalize_row_length(values, len(header))
        if _is_empty_row(normalized):
            continue
        total_rows += 1
        if len(reservoir) < max_points:
            reservoir.append(normalized)
        else:
            index = randomizer.randint(0, total_rows - 1)
            if index < max_points:
                reservoir[index] = normalized

    workbook.release_resources()
    if header is None:
        raise SystemExit(f"No header row found in {input_path}")

    frame = pl.DataFrame(reservoir, schema=header, orient="row", strict=False)
    sampling_note = None
    if total_rows > max_points:
        sampling_note = f"Reservoir sampling capped at {max_points:,} rows with seed {seed}"

    return PreparedFrame(
        frame=frame,
        total_rows=total_rows,
        sampled_rows=frame.height,
        sampling_note=sampling_note,
        source_name=input_path.name,
        sheet_name=sheet_name,
    )


def build_dashboard_payload(prepared: PreparedFrame, args: argparse.Namespace) -> dict[str, Any]:
    frame = _prepare_frame_for_dashboard(prepared.frame)
    if frame.height == 0:
        raise SystemExit("The sampled dataset has no data rows after cleaning.")

    column_meta = _build_column_meta(frame)
    defaults = _resolve_defaults(column_meta, args)
    return {
        "title": args.title or Path(prepared.source_name).stem.replace("_", " ").strip() or "3D Plot Dashboard",
        "sourceName": prepared.source_name,
        "sheetName": prepared.sheet_name,
        "rowCount": prepared.total_rows,
        "sampledRowCount": prepared.sampled_rows,
        "samplingNote": prepared.sampling_note,
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "columnOrder": frame.columns,
        "columnMeta": column_meta,
        "columns": {name: _serialize_series(frame.get_column(name), _meta_by_name(column_meta)[name]) for name in frame.columns},
        "defaults": defaults,
    }


def _prepare_frame_for_dashboard(frame: pl.DataFrame) -> pl.DataFrame:
    cleaned = _drop_empty_columns(frame)
    if cleaned.height == 0:
        return cleaned

    series_list: list[pl.Series] = []
    for name in cleaned.columns:
        series = cleaned.get_column(name)
        normalized = _normalize_series(name, series)
        if normalized is not None:
            series_list.append(normalized)

    prepared = pl.DataFrame(series_list)
    prepared = _append_derived_numeric_columns(prepared)
    return prepared


def _drop_empty_columns(frame: pl.DataFrame) -> pl.DataFrame:
    keep_columns: list[str] = []
    for name in frame.columns:
        series = frame.get_column(name)
        if series.dtype == pl.Null:
            continue
        if series.null_count() == frame.height:
            continue
        keep_columns.append(name)
    return frame.select(keep_columns)


def _normalize_series(name: str, series: pl.Series) -> pl.Series | None:
    if series.dtype.is_integer():
        return _downcast_integer_series(name, series)
    if series.dtype.is_float():
        return _downcast_float_series(name, series)
    if series.dtype == pl.Boolean:
        return series
    if series.dtype.is_temporal():
        return series.dt.strftime("%Y-%m-%d %H:%M:%S").alias(name)
    if series.dtype == pl.Object:
        values = [_normalize_python_value(value) for value in series.to_list()]
        normalized = pl.Series(name, values, strict=False)
        if normalized.dtype == pl.Object:
            return normalized.cast(pl.String, strict=False).alias(name)
        return normalized.alias(name)
    return series.cast(pl.String, strict=False).alias(name)


def _downcast_integer_series(name: str, series: pl.Series) -> pl.Series:
    non_null = series.drop_nulls()
    if non_null.len() == 0:
        return series.cast(pl.Int32, strict=False).alias(name)
    minimum = int(non_null.min())
    maximum = int(non_null.max())
    for dtype, low, high in (
        (pl.UInt8, 0, 255),
        (pl.UInt16, 0, 65_535),
        (pl.UInt32, 0, 4_294_967_295),
        (pl.Int8, -128, 127),
        (pl.Int16, -32_768, 32_767),
        (pl.Int32, -2_147_483_648, 2_147_483_647),
    ):
        if minimum >= low and maximum <= high:
            return series.cast(dtype, strict=False).alias(name)
    return series.cast(pl.Int64, strict=False).alias(name)


def _downcast_float_series(name: str, series: pl.Series) -> pl.Series:
    return series.cast(pl.Float32, strict=False).alias(name)


def _append_derived_numeric_columns(frame: pl.DataFrame) -> pl.DataFrame:
    if frame.height == 0:
        return frame

    numeric_columns = [name for name in frame.columns if frame.get_column(name).dtype.is_numeric()]
    derived_columns: list[pl.Series] = []

    if len(numeric_columns) < 3:
        derived_columns.append(pl.Series("__row_index", list(range(frame.height)), dtype=pl.Int32))

    categorical_candidates = [
        name
        for name in frame.columns
        if not frame.get_column(name).dtype.is_numeric() and name != "__row_index"
    ]
    for name in categorical_candidates:
        if len(numeric_columns) + len(derived_columns) >= 3:
            break
        encoded = _factorize_to_ints(frame.get_column(name))
        derived_columns.append(pl.Series(f"__code_{name}", encoded, dtype=pl.Int32))

    if derived_columns:
        frame = frame.hstack(derived_columns)

    numeric_after = [name for name in frame.columns if frame.get_column(name).dtype.is_numeric()]
    if len(numeric_after) < 3:
        raise SystemExit(
            "At least three numeric dimensions are required after preprocessing. "
            "The input did not contain enough usable values to derive them."
        )
    return frame


def _build_column_meta(frame: pl.DataFrame) -> list[dict[str, Any]]:
    meta: list[dict[str, Any]] = []
    for name in frame.columns:
        series = frame.get_column(name)
        column_kind = _classify_column(series)
        item: dict[str, Any] = {
            "name": name,
            "dtype": str(series.dtype),
            "kind": column_kind,
            "derived": name.startswith("__"),
            "nullCount": series.null_count(),
        }
        if column_kind == "numeric":
            non_null = series.drop_nulls()
            item["min"] = _safe_number(non_null.min()) if non_null.len() else None
            item["max"] = _safe_number(non_null.max()) if non_null.len() else None
        else:
            unique_count = int(series.n_unique()) if series.len() else 0
            item["uniqueCount"] = unique_count
        meta.append(item)
    return meta


def _resolve_defaults(column_meta: list[dict[str, Any]], args: argparse.Namespace) -> dict[str, Any]:
    meta_by_name = _meta_by_name(column_meta)
    numeric = [item["name"] for item in column_meta if item["kind"] == "numeric"]
    preferred_numeric = [name for name in numeric if not meta_by_name[name]["derived"]] + [
        name for name in numeric if meta_by_name[name]["derived"]
    ]
    categorical = [item["name"] for item in column_meta if item["kind"] in {"categorical", "boolean"}]
    labels = [item["name"] for item in column_meta if item["kind"] != "numeric"]

    def choose_axis(requested: str | None, fallback: str) -> str:
        if requested is None:
            return fallback
        _require_column(meta_by_name, requested, allowed_kinds={"numeric"})
        return requested

    x_name = choose_axis(args.x, preferred_numeric[0])
    y_name = choose_axis(args.y, next(name for name in preferred_numeric if name != x_name))
    z_name = choose_axis(args.z, next(name for name in preferred_numeric if name not in {x_name, y_name}))

    if args.color:
        _require_column(meta_by_name, args.color)
        color_name = args.color
    else:
        color_name = next((name for name in categorical if name not in {x_name, y_name, z_name}), None)
        if color_name is None:
            color_name = next((name for name in preferred_numeric if name not in {x_name, y_name, z_name}), None)

    if args.size:
        _require_column(meta_by_name, args.size, allowed_kinds={"numeric"})
        size_name = args.size
    else:
        size_name = next((name for name in preferred_numeric if name not in {x_name, y_name, z_name}), None)

    if args.label:
        _require_column(meta_by_name, args.label)
        label_name = args.label
    else:
        label_name = next((name for name in labels if not meta_by_name[name]["derived"]), None)

    filter_column = next((name for name in categorical if not meta_by_name[name]["derived"]), None)

    return {
        "x": x_name,
        "y": y_name,
        "z": z_name,
        "color": color_name,
        "size": size_name,
        "label": label_name,
        "filterColumn": filter_column,
    }


def _serialize_series(series: pl.Series, meta: dict[str, Any]) -> dict[str, Any]:
    if meta["kind"] == "numeric":
        return {
            "encoding": "number",
            "data": [_safe_number(value) for value in series.to_list()],
        }

    if meta["kind"] in {"categorical", "boolean"}:
        vocab, codes = _dictionary_encode(series)
        return {
            "encoding": "dictionary",
            "values": vocab,
            "codes": codes,
        }

    return {
        "encoding": "string",
        "data": [_string_or_none(value) for value in series.to_list()],
    }


def _classify_column(series: pl.Series) -> str:
    if series.dtype.is_numeric():
        return "numeric"
    if series.dtype == pl.Boolean:
        return "boolean"

    unique_count = int(series.n_unique()) if series.len() else 0
    threshold = max(32, min(512, int(series.len() * 0.18)))
    if unique_count <= threshold:
        return "categorical"
    return "text"


def _dictionary_encode(series: pl.Series) -> tuple[list[Any], list[int | None]]:
    vocabulary: list[Any] = []
    index_by_value: dict[Any, int] = {}
    codes: list[int | None] = []
    for value in series.to_list():
        normalized = _normalize_jsonish_value(value)
        if normalized is None:
            codes.append(None)
            continue
        if normalized not in index_by_value:
            index_by_value[normalized] = len(vocabulary)
            vocabulary.append(normalized)
        codes.append(index_by_value[normalized])
    return vocabulary, codes


def _factorize_to_ints(series: pl.Series) -> list[int | None]:
    values, codes = _dictionary_encode(series)
    if not values:
        return [None] * series.len()
    return codes


def _meta_by_name(column_meta: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {item["name"]: item for item in column_meta}


def _require_column(
    meta_by_name: dict[str, dict[str, Any]],
    name: str,
    allowed_kinds: set[str] | None = None,
) -> None:
    if name not in meta_by_name:
        raise SystemExit(f"Unknown column: {name}")
    if allowed_kinds and meta_by_name[name]["kind"] not in allowed_kinds:
        kinds = ", ".join(sorted(allowed_kinds))
        raise SystemExit(f"Column '{name}' must be one of: {kinds}")


def _normalize_headers(values: Sequence[Any]) -> list[str]:
    counts: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values):
        base = str(value).strip() if value is not None and str(value).strip() else f"column_{index + 1}"
        count = counts.get(base, 0)
        headers.append(base if count == 0 else f"{base}_{count + 1}")
        counts[base] = count + 1
    return headers


def _normalize_row_length(values: Sequence[Any], width: int) -> list[Any]:
    normalized = list(values[:width])
    if len(normalized) < width:
        normalized.extend([None] * (width - len(normalized)))
    return normalized


def _normalize_python_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time())
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return value


def _normalize_xls_cell(cell: xlrd.sheet.Cell, datemode: int) -> Any:
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return int(cell.value) if float(cell.value).is_integer() else float(cell.value)
    if cell.ctype == xlrd.XL_CELL_DATE:
        year, month, day, hour, minute, second = xlrd.xldate_as_tuple(cell.value, datemode)
        if (year, month, day) == (0, 0, 0):
            return f"{hour:02d}:{minute:02d}:{second:02d}"
        return datetime(year, month, day, hour, minute, second)
    return cell.value


def _is_empty_row(values: Sequence[Any]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


def _select_openpyxl_sheet(workbook: Any, requested_sheet: str | None) -> tuple[str, Any]:
    sheet_names = workbook.sheetnames
    if not sheet_names:
        raise SystemExit("The workbook does not contain any sheets.")
    if requested_sheet is None:
        selected = sheet_names[0]
        return selected, workbook[selected]
    if requested_sheet.isdigit():
        index = int(requested_sheet)
        if not 0 <= index < len(sheet_names):
            raise SystemExit(f"Sheet index out of range: {requested_sheet}")
        selected = sheet_names[index]
        return selected, workbook[selected]
    if requested_sheet not in sheet_names:
        available = ", ".join(sheet_names)
        raise SystemExit(f"Sheet '{requested_sheet}' not found. Available sheets: {available}")
    return requested_sheet, workbook[requested_sheet]


def _select_xlrd_sheet(workbook: xlrd.book.Book, requested_sheet: str | None) -> tuple[str, xlrd.sheet.Sheet]:
    sheet_names = workbook.sheet_names()
    if not sheet_names:
        raise SystemExit("The workbook does not contain any sheets.")
    if requested_sheet is None:
        return sheet_names[0], workbook.sheet_by_index(0)
    if requested_sheet.isdigit():
        index = int(requested_sheet)
        if not 0 <= index < len(sheet_names):
            raise SystemExit(f"Sheet index out of range: {requested_sheet}")
        return sheet_names[index], workbook.sheet_by_index(index)
    if requested_sheet not in sheet_names:
        available = ", ".join(sheet_names)
        raise SystemExit(f"Sheet '{requested_sheet}' not found. Available sheets: {available}")
    return requested_sheet, workbook.sheet_by_name(requested_sheet)


def _detect_separator(input_path: Path) -> str:
    if input_path.suffix.lower() == ".tsv":
        return "\t"
    sample = input_path.read_text(encoding="utf-8", errors="ignore")[:8_192]
    first_lines = [line for line in sample.splitlines() if line.strip()][:5]
    if not first_lines:
        return ","
    scores = {candidate: sum(line.count(candidate) for line in first_lines) for candidate in [",", "\t", ";", "|"]}
    return max(scores, key=scores.get)


def _safe_number(value: Any) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(numeric) or math.isinf(numeric):
        return None
    return int(numeric) if numeric.is_integer() else numeric


def _string_or_none(value: Any) -> str | None:
    normalized = _normalize_jsonish_value(value)
    if normalized is None:
        return None
    return str(normalized)


def _normalize_jsonish_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, Decimal):
        return float(value)
    return value
